import time, sys, urllib.request, json, urllib.error, csv
from datetime import datetime, timezone

# Raw listings path within any SimplifyJobs repo
LISTINGS_PATH = "dev/.github/scripts/listings.json"
RAW_BASE = "https://raw.githubusercontent.com/SimplifyJobs/{repo}/" + LISTINGS_PATH
NEWGRAD_REPO = "New-Grad-Positions"
INTERN_REPO_TMPL = "Summer{year}-Internships"

# Fallback if auto-detection can't reach GitHub (kept current-ish; auto-detect
# normally overrides this anyway).
FALLBACK_INTERN_REPO = "Summer2026-Internships"

# Category labels as they appear in the feed
SOFTWARE_CATEGORIES = {"Software", "Software Engineering"}
AI_CATEGORIES = {"AI/ML/Data", "Data Science, AI & Machine Learning"}

# Canadian province / territory codes (used as a secondary location signal)
CA_PROVINCES = {
    "ON", "QC", "BC", "AB", "MB", "SK", "NS", "NB", "NL", "PE", "NT", "YT", "NU",
}

# Title must look like an internship or a (Canadian) co-op role
INTERN_KEYWORDS = ("intern", "co-op", "coop", "co op")

# Software-adjacent role types. Used by --rescue-adjacent to catch roles that
# the feed mis-categorised (e.g. a frontend role tagged "AI/ML/Data").
SWE_ADJACENT_KEYWORDS = (
    "software", "developer", "programmer", "swe",
    "front end", "front-end", "frontend",
    "back end", "back-end", "backend",
    "full stack", "full-stack", "fullstack",
    "devops", "dev ops", "sre", "site reliability",
    "web develop", "mobile", "ios", "android",
    "cloud engineer", "platform engineer", "infrastructure engineer",
    "sdet", "qa engineer", "automation engineer",
)

# When rescuing, skip these to avoid pulling in hardware/firmware roles.
HARDWARE_EXCLUDE = ("embedded", "firmware", "fpga", "circuit", "pcb", "rtl", "asic")

# caches
_cache = {}              # url -> (fetched_at, listings)
CACHE_TTL = 300          # matches GitHub's own cache header (5 minutes)
_url_cache = {}          # (source, year, source_url) -> (resolved_at, url)
URL_TTL = 3600           # the active repo barely changes (1 hour)


def get_jobs(search_parameters):
    url = get_source_url(search_parameters)
    try:
        listings = get_listings(url)
    except Exception as e:
        raise Exception(f"could not download the feed: {e}")

    jobs = find_jobs(listings,search_parameters)
    return jobs

def find_jobs(listings, search_parameters):
    now = time.time()
    categories = select_categories(search_parameters)
    roles = None

    if search_parameters.roles:
        roles = [r.strip().lower() for r in search_parameters.roles.split(",") if r.strip()]
    term = search_parameters.term.strip().lower() if search_parameters.term else None
    window = search_parameters.hours * 3600
    jobs_output = []

    for job in listings:
        if not job.get("active", False):
            continue
        if not job.get("is_visible", True):
            continue
        title = job.get("title", "")

        # Season / term filter (e.g. "fall", "winter 2027"). Feeds without a
        # terms field (e.g. New-Grad) simply skip this filter.
        if term:
            job_terms = " ".join(job.get("terms") or []).lower()
            if job_terms and term not in job_terms:
                continue
            if not job_terms:  # no term info but user asked for a season
                continue

        # Category gate, with optional rescue of mis-categorised SWE-adjacent roles
        in_category = categories is None or job.get("category") in categories
        if not in_category:
            if search_parameters.rescue_adjacent and is_swe_adjacent(title):
                pass  # rescued by title
            else:
                continue

        if search_parameters.source != "newgrad" and not is_internship(title):
            continue

        # Optional narrowing to specific role keywords (e.g. --roles devops,backend)
        if roles and not any(r in title.lower() for r in roles):
            continue

        if not is_in_canada(job.get("locations"), search_parameters.province):
            continue

        posted = job.get("date_posted", 0)
        if (now - posted) > window:
            continue
        jobs_output.append(job)

    jobs_output.sort(key=lambda j: j.get("date_posted", 0), reverse=True)

    return jobs_output

def select_categories(search_parameters):
    cats = set(SOFTWARE_CATEGORIES)
    if search_parameters.include_ai:
        cats |= AI_CATEGORIES
    if search_parameters.all_tech:
        return None  # None => accept every category
    return cats

def is_in_canada(locations, province=None):
    """True if any location string points to Canada (optionally one province)."""
    for loc in locations or []:
        low = loc.lower()
        canadian = ("canada" in low) or loc.strip().endswith(", CAN")
        if not canadian:
            # also catch bare 'City, ON' style with a province code + no country
            parts = [p.strip() for p in loc.split(",")]
            if any(p in CA_PROVINCES for p in parts):
                canadian = True
        if not canadian:
            continue
        if province:
            parts = [p.strip().upper() for p in loc.split(",")]
            if province.upper() not in parts:
                continue
        return True
    return False

def is_internship(title):
    t = (title or "").lower()
    return any(k in t for k in INTERN_KEYWORDS)

def is_swe_adjacent(title):
    """True if the title names a software-adjacent role (frontend, devops, ...)."""
    t = (title or "").lower()
    if any(h in t for h in HARDWARE_EXCLUDE):
        return False
    return any(k in t for k in SWE_ADJACENT_KEYWORDS)


def get_listings(url):
    now = time.monotonic() #  monotonic clock is a clock that always moves forward and is unaffected by system clock changes
    hit = _cache.get(url)
    if hit and now - hit[0] < CACHE_TTL:
        # print("CACHE HIT - reusing")   
        return hit[1]                 # fresh -> reuse, no network
    # print("CACHE MISS - fetching from network")
    listings = fetch_listings(url)    # stale or missing -> fetch
    _cache[url] = (now, listings)     
    return listings

def fetch_listings(url):
    """Download and parse a listings feed from a given URL."""
    req = urllib.request.Request(url, headers={"User-Agent": "intern-finder/1.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode("utf-8"))

def get_source_url(search_parameters):
    key = (search_parameters.source, search_parameters.year, search_parameters.source_url)
    now = time.monotonic()
    hit = _url_cache.get(key)
    if hit and now - hit[0] < URL_TTL:
        return hit[1]
    url = resolve_source_url(search_parameters)
    _url_cache[key] = (now, url)
    return url

def resolve_source_url(search_parameters, verbose=True):
    """Decide which listings.json URL to pull from, based on search_parameters."""
    if search_parameters.source_url:
        return search_parameters.source_url
    if search_parameters.source == "newgrad":
        # if verbose:
        #     print(f"  Source: {NEWGRAD_REPO} (full-time new-grad roles)")
        return RAW_BASE.format(repo=NEWGRAD_REPO)
    # default: internships
    if search_parameters.year:
        repo = INTERN_REPO_TMPL.format(year=search_parameters.year)
    else:
        repo = detect_intern_repo(verbose=verbose)
    return RAW_BASE.format(repo=repo)

def detect_intern_repo(verbose=True):
    """Find the newest existing 'Summer{YEAR}-Internships' repo.

    SimplifyJobs launches a new repo for each upcoming cycle, so the highest
    year that exists is the active one. This makes the script self-updating
    across years. Falls back to a known repo if GitHub can't be reached.
    """
    this_year = datetime.now(timezone.utc).year
    # check next two years down to last year (covers early/late-cycle naming)
    candidates = [this_year + 1, this_year, this_year - 1]
    reachable_any = False
    for y in candidates:
        repo = INTERN_REPO_TMPL.format(year=y)
        exists = _listings_exists(repo)
        if exists is not None:
            reachable_any = True
        if exists:
            if verbose:
                print(f"  Active internship repo detected: {repo}")
            return repo
    if not reachable_any and verbose:
        print("  (Could not reach GitHub; using fallback repo.)")
    elif verbose:
        print(f"  Using fallback repo: {FALLBACK_INTERN_REPO}")
    return FALLBACK_INTERN_REPO

def _listings_exists(repo, timeout=15):
    """Cheap existence check: a 1-byte Range request to the raw listings file.

    Avoids the GitHub API's 60/hour unauthenticated rate limit, and checks the
    exact file we need rather than just the repo.
    """
    url = RAW_BASE.format(repo=repo)
    req = urllib.request.Request(
        url, headers={"User-Agent": "intern-finder/1.0", "Range": "bytes=0-0"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.status in (200, 206)
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return False
        return None  # other HTTP error -> uncertain
    except Exception:
        return None  # network/uncertain -> caller decides


# export functions

def export_csv(jobs, path, now):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            ["title", "company", "category", "locations",
             "posted_utc", "hours_ago", "sponsorship", "url"]
        )
        for j in jobs:
            w.writerow([
                j.get("title", ""),
                j.get("company_name", ""),
                j.get("category", ""),
                "; ".join(j.get("locations") or []),
                datetime.fromtimestamp(
                    j.get("date_posted", 0), timezone.utc
                ).strftime("%Y-%m-%d %H:%M"),
                f"{hours_since(j.get('date_posted',0), now):.1f}",
                j.get("sponsorship", ""),
                j.get("url", ""),
            ])
    print(f"\nSaved {len(jobs)} role(s) to {path}")


def export_xlsx(jobs, path, now):
    """Write results to a formatted .xlsx workbook.

    Columns: Job Title | Company | Location | Category | Date Posted |
             Deadline to Apply | Description & Application Link

    Note: the upstream feed does NOT publish application deadlines, so that
    column is flagged for manual checking rather than left misleadingly blank.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.comments import Comment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl is required for Excel export. Install it with:\n"
              "    pip install openpyxl", file=sys.stderr)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "SWE Interns Canada"

    headers = ["Job Title", "Company", "Location", "Category",
               "Date Posted", "Deadline to Apply",
               "Description & Application Link"]
    ws.append(headers)

    header_fill = PatternFill("solid", start_color="1F4E78")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.cell(row=1, column=6).comment = Comment(
        "The internship feed does not publish deadlines. Open the posting "
        "link to confirm the closing date — many internships are rolling / "
        "'until filled'.", "intern-finder")

    body_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    deadline_fill = PatternFill("solid", start_color="FFF2CC")

    for j in jobs:
        posted = datetime.fromtimestamp(
            j.get("date_posted", 0), timezone.utc).date()
        url = j.get("url", "")
        ws.append([
            j.get("title", ""),
            j.get("company_name", ""),
            ", ".join(j.get("locations") or []),
            j.get("category", ""),
            posted,
            "Not listed — check posting",
            url,
        ])
        r = ws.max_row
        for col in range(1, 8):
            cell = ws.cell(row=r, column=col)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (1, 3)))
        ws.cell(row=r, column=5).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=6).fill = deadline_fill
        link_cell = ws.cell(row=r, column=7)
        if url:
            link_cell.hyperlink = url
            link_cell.font = link_font

    widths = [40, 22, 30, 18, 13, 24, 55]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    wb.save(path)
    print(f"\nSaved {len(jobs)} role(s) to {path}")

def hours_since(epoch, now):
    return (now - epoch) / 3600.0
