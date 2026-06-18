#!/usr/bin/env python3

import argparse
import csv
import json
import sys
import time
from datetime import datetime, timezone
from core import select_categories, resolve_source_url,is_in_canada, is_internship, is_swe_adjacent, fetch_listings


def hours_since(epoch, now):
    return (now - epoch) / 3600.0


def find_jobs_cli_version(listings, args, now=None):
    now = now or time.time()
    cats = select_categories(args)
    roles = None
    if args.roles:
        roles = [r.strip().lower() for r in args.roles.split(",") if r.strip()]
    term = args.term.strip().lower() if args.term else None
    window = args.hours * 3600
    out = []
    for job in listings:
        if not job.get("active", False):
            continue
        if not job.get("is_visible", True):
            continue
        title = job.get("title", "")

        # Season / term filter (e.g. "fall", "winter 2027"). Feeds without a
        # terms field (e.g. New-Grad) simply skip this filter.
        if term:
            job_terms = " ".join(job.get("terms") or []).lower()
            if job_terms and term not in job_terms:
                continue
            if not job_terms:  # no term info but user asked for a season
                continue

        # Category gate, with optional rescue of mis-categorised SWE-adjacent roles
        in_category = cats is None or job.get("category") in cats
        if not in_category:
            if args.rescue_adjacent and is_swe_adjacent(title):
                pass  # rescued by title
            else:
                continue

        if args.source != "newgrad" and not is_internship(title):
            continue
        # Optional narrowing to specific role keywords (e.g. --roles devops,backend)
        if roles and not any(r in title.lower() for r in roles):
            continue
        if not is_in_canada(job.get("locations"), args.province):
            continue
        posted = job.get("date_posted", 0)
        if (now - posted) > window:
            continue
        out.append(job)
    out.sort(key=lambda j: j.get("date_posted", 0), reverse=True)
    return out


def print_jobs(jobs, args, now):
    if not jobs:
        print(
            f"\nNo matching roles posted in the last {args.hours}h.\n"
            "Tip: summer-internship posting peaks Aug–Jan, so a 24h window is\n"
            "often empty in late spring/summer. Try --hours 72, --hours 168,\n"
            "--include-ai, or --all-tech to widen the search."
        )
        return
    print(f"\nFound {len(jobs)} role(s) posted in the last {args.hours}h:\n")
    for j in jobs:
        age = hours_since(j.get("date_posted", 0), now)
        when = datetime.fromtimestamp(
            j.get("date_posted", 0), timezone.utc
        ).strftime("%Y-%m-%d %H:%M UTC")
        print(f"{'='*70}")
        print(f"  {j.get('title','(no title)')}")
        print(f"  {j.get('company_name','(unknown)')}  |  {j.get('category','')}")
        print(f"  Location : {', '.join(j.get('locations') or []) or 'n/a'}")
        print(f"  Posted   : {when}  ({age:.1f}h ago)")
        if j.get("sponsorship"):
            print(f"  Sponsor. : {j.get('sponsorship')}")
        print(f"  Apply    : {j.get('url','')}")
    print(f"{'='*70}")


def export_csv(jobs, path, now):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            ["title", "company", "category", "locations",
             "posted_utc", "hours_ago", "sponsorship", "url"]
        )
        for j in jobs:
            w.writerow([
                j.get("title", ""),
                j.get("company_name", ""),
                j.get("category", ""),
                "; ".join(j.get("locations") or []),
                datetime.fromtimestamp(
                    j.get("date_posted", 0), timezone.utc
                ).strftime("%Y-%m-%d %H:%M"),
                f"{hours_since(j.get('date_posted',0), now):.1f}",
                j.get("sponsorship", ""),
                j.get("url", ""),
            ])
    print(f"\nSaved {len(jobs)} role(s) to {path}")


def export_xlsx(jobs, path, now):
    """Write results to a formatted .xlsx workbook.

    Columns: Job Title | Company | Location | Category | Date Posted |
             Deadline to Apply | Description & Application Link

    Note: the upstream feed does NOT publish application deadlines, so that
    column is flagged for manual checking rather than left misleadingly blank.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.comments import Comment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl is required for Excel export. Install it with:\n"
              "    pip install openpyxl", file=sys.stderr)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "SWE Interns Canada"

    headers = ["Job Title", "Company", "Location", "Category",
               "Date Posted", "Deadline to Apply",
               "Description & Application Link"]
    ws.append(headers)

    header_fill = PatternFill("solid", start_color="1F4E78")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.cell(row=1, column=6).comment = Comment(
        "The internship feed does not publish deadlines. Open the posting "
        "link to confirm the closing date — many internships are rolling / "
        "'until filled'.", "intern-finder")

    body_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    deadline_fill = PatternFill("solid", start_color="FFF2CC")

    for j in jobs:
        posted = datetime.fromtimestamp(
            j.get("date_posted", 0), timezone.utc).date()
        url = j.get("url", "")
        ws.append([
            j.get("title", ""),
            j.get("company_name", ""),
            ", ".join(j.get("locations") or []),
            j.get("category", ""),
            posted,
            "Not listed — check posting",
            url,
        ])
        r = ws.max_row
        for col in range(1, 8):
            cell = ws.cell(row=r, column=col)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (1, 3)))
        ws.cell(row=r, column=5).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=6).fill = deadline_fill
        link_cell = ws.cell(row=r, column=7)
        if url:
            link_cell.hyperlink = url
            link_cell.font = link_font

    widths = [40, 22, 30, 18, 13, 24, 55]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    wb.save(path)
    print(f"\nSaved {len(jobs)} role(s) to {path}")


def main(argv=None):
    p = argparse.ArgumentParser(
        description="Find recently-posted SWE internships/co-ops in Canada."
    )
    p.add_argument("--hours", type=float, default=24,
                   help="Only roles posted within this many hours (default 24).")
    p.add_argument("--term", default=None,
                   help="Season filter, e.g. 'fall', 'winter', 'spring', "
                        "'summer', or a specific one like 'winter 2027'. "
                        "Default: all terms.")
    p.add_argument("--source", choices=["internships", "newgrad"],
                   default="internships",
                   help="Which feed to use: internships (default) or newgrad "
                        "(full-time new-grad roles).")
    p.add_argument("--year", type=int, default=None,
                   help="Force a specific internship cycle year (e.g. 2027). "
                        "By default the active repo is auto-detected.")
    p.add_argument("--source-url", default=None,
                   help="Override with a custom listings.json URL.")
    p.add_argument("--province", default=None,
                   help="Restrict to one province code, e.g. ON, QC, BC.")
    p.add_argument("--include-ai", action="store_true",
                   help="Also include AI/ML/Data roles.")
    p.add_argument("--all-tech", action="store_true",
                   help="Include every tech category (hardware, quant, PM...).")
    p.add_argument("--rescue-adjacent", action="store_true",
                   help="Also catch frontend/backend/devops/cloud/SRE roles even "
                        "when the feed mis-tags them under AI/ML or other categories "
                        "(hardware/firmware titles are still excluded).")
    p.add_argument("--roles", metavar="KW", default=None,
                   help="Comma-separated title keywords to narrow results, "
                        "e.g. --roles devops,backend,frontend")
    p.add_argument("--csv", metavar="PATH", default=None,
                   help="Also write results to a CSV file.")
    p.add_argument("--xlsx", metavar="PATH", default=None,
                   help="Also write results to a formatted Excel (.xlsx) file.")
    p.add_argument("--json", metavar="PATH", default=None,
                   help="Also write raw matching entries to a JSON file.")
    args = p.parse_args(argv)

    print("Resolving live internship feed (SimplifyJobs / Pitt CSC)...")
    url = resolve_source_url(args)
    try:
        listings = fetch_listings(url)
    except Exception as e:
        print(f"ERROR: could not download the feed: {e}", file=sys.stderr)
        return 1

    now = time.time()
    jobs = find_jobs_cli_version(listings, args, now=now)
    print_jobs(jobs, args, now)

    if args.csv and jobs:
        export_csv(jobs, args.csv, now)
    if args.xlsx and jobs:
        export_xlsx(jobs, args.xlsx, now)
    if args.json and jobs:
        with open(args.json, "w", encoding="utf-8") as f:
            json.dump(jobs, f, indent=2)
        print(f"Saved raw JSON to {args.json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())